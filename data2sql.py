import os,sys,django
os.environ.setdefault('DJANGO_SETTINGS_MODULE',"onlineproject.settings")
django.setup()

import click
from bs4 import BeautifulSoup
from openpyxl import *
from onlineapp.models import *

@click.group()
def cli():
    pass


@cli.command("createdb",help = "to create student, college and marks table")
def createdb():
    pass


@cli.command("dropdb",help="to drop dbs")
def dropdb():
    College.objects.all().delete()


@cli.command("populatedb",help="to fill the db with values from .xlsx sheets and html files")
def populatedb():
    details = load_workbook("students.xlsx")
    colleges = details["Colleges"]
    for row in range(2, colleges.max_row + 1):
        insert = []
        for col in range(1, colleges.max_column + 1):
            insert.append(colleges.cell(row=row, column=col).value)
        try:
            c = College(name=insert[0], location=insert[2], acronym=insert[1], contact=insert[3])
            c.save()
            #print(c)
        except Exception as e:
            pass
    students = details["Current"]
    for row in range(2, students.max_row + 1):
        insert = []
        for col in range(1, students.max_column + 1):
            insert.append(students.cell(row=row, column=col).value)
        try:
            s = Student(name=insert[0], email=insert[2], college=College.objects.get(acronym=insert[1]),
                        db_folder=insert[3].lower())
            s.save()
        except Exception as e:
            pass

    students = details["Deletions"]
    for row in range(2, students.max_row + 1):
        insert = []
        for col in range(1, students.max_column + 1):
            insert.append(students.cell(row=row, column=col).value)
        try:
            s = Student(name=insert[0], email=insert[2], college=College.objects.get(acronym=insert[1]),
                        db_folder=insert[3].lower(),dropped_out=True)
            s.save()
        except Exception as e:
            pass

    file = open("mock_results.html")
    soup = BeautifulSoup(file, "html.parser")
    table_rows = soup.find_all('tr')
    row_no, column_no = 2, 1
    for table_row in table_rows[1:]:
        insert = []
        for row_data in table_row.find_all('td')[1:]:
            insert.append(row_data.text)
            column_no += 1
        column_no = 1
        x = insert[0].split('_')
        try:
            m = MockTest1(student=Student.objects.get(db_folder=x[2].lower()), problem1=insert[1], problem2=insert[2],
                          problem3=insert[3], problem4=insert[4], total=insert[5])
            m.save()
        except Exception as e:
            print(e)
        row_no += 1

    print("inserted into database")


if __name__ == '__main__':
    cli()
